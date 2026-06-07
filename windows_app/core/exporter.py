"""
Export des résultats vers CSV / Excel.
Utilise csv (stdlib) + openpyxl directement, sans pandas/numpy (build .exe plus fiable).
"""
import os
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


COLUMNS = [
    "URL", "URL finale", "Nom Entreprise", "E-commerce", "Verdict 3DS", "Score 3DS",
    "Passerelles détectées", "Mots-clés 3DS", "Emails", "Téléphones",
    "Formulaire contact", "Pages contact", "Facebook", "Instagram", "LinkedIn",
    "Raison", "Status HTTP",
]


def _rows_from_results(results):
    rows = []
    for r in results:
        contacts = r.get("contacts") or {}
        gw_names = ", ".join(g["name"] for g in r.get("gateways", []))
        rows.append({
            "URL": r.get("url", ""),
            "URL finale": r.get("final_url", ""),
            "Nom Entreprise": contacts.get("company_name", "") or "",
            "E-commerce": "Oui" if r.get("is_ecommerce") else "Non",
            "Verdict 3DS": r.get("verdict", ""),
            "Score 3DS": r.get("score_3ds", 0),
            "Passerelles détectées": gw_names,
            "Mots-clés 3DS": ", ".join(r.get("keywords_found", [])),
            "Emails": ", ".join(contacts.get("emails", [])),
            "Téléphones": ", ".join(contacts.get("phones", [])),
            "Formulaire contact": "Oui" if contacts.get("has_contact_form") else "Non",
            "Pages contact": ", ".join(contacts.get("contact_pages", [])),
            "Facebook": contacts.get("socials", {}).get("facebook", ""),
            "Instagram": contacts.get("socials", {}).get("instagram", ""),
            "LinkedIn": contacts.get("socials", {}).get("linkedin", ""),
            "Raison": r.get("reason", ""),
            "Status HTTP": r.get("status", "") if r.get("status") is not None else "",
        })
    return rows


def export_csv(results, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    rows = _rows_from_results(results)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"prospects_3ds_{ts}.csv")

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, delimiter=";")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return path


def _write_sheet(ws, rows, columns):
    """Écrit une feuille avec entêtes stylisées + auto-fit."""
    # Header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="00B894")
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit columns
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row in rows:
            val = str(row.get(col_name, ""))
            if len(val) > max_len:
                max_len = len(val)
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # Freeze header
    ws.freeze_panes = "A2"


def export_excel(results, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    rows = _rows_from_results(results)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"prospects_3ds_{ts}.xlsx")

    wb = Workbook()
    # Feuille 1 : Tous
    ws_all = wb.active
    ws_all.title = "Tous les sites"
    _write_sheet(ws_all, rows, COLUMNS)

    # Feuille 2 : Prospects (sans 3DS ou Incertain)
    prospects = [r for r in rows if r["Verdict 3DS"] in ("Sans 3DS Probable", "Incertain")]
    if prospects:
        ws_p = wb.create_sheet("Prospects")
        _write_sheet(ws_p, prospects, COLUMNS)

    # Feuille 3 : Déjà 3DS
    ok = [r for r in rows if r["Verdict 3DS"] == "3DS Probable"]
    if ok:
        ws_ok = wb.create_sheet("Déjà 3DS")
        _write_sheet(ws_ok, ok, COLUMNS)

    wb.save(path)
    return path
